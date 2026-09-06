#!/usr/bin/env python3
"""Product Creator for the Digital Storefront Planner.

Generates digital products (planners, SVGs, spreadsheets, social templates,
wall art, resumes, checklists) with mockup images and metadata.
"""

import os
import sys
import json
import logging
import argparse
import math
import shutil
import tempfile
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Any, Optional

from fpdf import FPDF
from PIL import Image, ImageDraw, ImageFont, ImageFilter
import svgwrite
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
RESOURCES_DIR = SCRIPT_DIR / "resources"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR.parent / "output"

logger = logging.getLogger("product_creator")

STYLE_PALETTES = {
    "minimalist": {
        "primary": "#2D3436",
        "secondary": "#636E72",
        "accent": "#00B894",
        "background": "#FFFFFF",
        "text": "#2D3436",
    },
    "modern": {
        "primary": "#6C5CE7",
        "secondary": "#A29BFE",
        "accent": "#FD79A8",
        "background": "#FAFAFA",
        "text": "#2D3436",
    },
    "colorful": {
        "primary": "#E17055",
        "secondary": "#FDCB6E",
        "accent": "#00CEC9",
        "background": "#FFFFFF",
        "text": "#2D3436",
    },
    "botanical": {
        "primary": "#2D6A4F",
        "secondary": "#52B788",
        "accent": "#95D5B2",
        "background": "#F8F9FA",
        "text": "#1B4332",
    },
    "geometric": {
        "primary": "#003049",
        "secondary": "#D62828",
        "accent": "#F77F00",
        "background": "#FCBF49",
        "text": "#003049",
    },
}


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------


@dataclass
class ProductConcept:
    """Describes a digital product to be generated."""

    name: str
    slug: str
    category: str
    description: str
    features: list[str] = field(default_factory=list)
    style: str = "minimalist"
    colors: dict = field(default_factory=dict)
    dimensions: dict = field(
        default_factory=lambda: {"width": 612, "height": 792, "unit": "pt"}
    )
    pages: int = 5
    target_audience: str = "general"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    """Convert a hex colour string (#RRGGBB or RRGGBB) to an (R, G, B) tuple."""
    hex_color = hex_color.lstrip("#")
    return (
        int(hex_color[0:2], 16),
        int(hex_color[2:4], 16),
        int(hex_color[4:6], 16),
    )


# ---------------------------------------------------------------------------
# ProductCreator
# ---------------------------------------------------------------------------


class ProductCreator:
    """Factory that dispatches product generation by category."""

    CATEGORY_MAP = {
        "printable_planner": "create_printable_planner",
        "svg_design": "create_svg_design",
        "spreadsheet_template": "create_spreadsheet",
        "social_media_template": "create_social_template",
        "wall_art": "create_wall_art",
        "resume_template": "create_resume_template",
        "checklist": "create_checklist",
        "digital_sticker": "create_social_template",
    }

    def __init__(
        self,
        output_dir: Optional[Path] = None,
        capabilities_path: Optional[Path] = None,
    ):
        self.output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.capabilities = self._load_capabilities(capabilities_path)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _load_capabilities(self, path: Optional[Path]) -> dict:
        """Load the capabilities JSON if it exists, else return defaults."""
        if path and Path(path).is_file():
            with open(path, "r") as fh:
                return json.load(fh)
        default_path = RESOURCES_DIR / "capabilities.json"
        if default_path.is_file():
            with open(default_path, "r") as fh:
                return json.load(fh)
        return {
            "product_types": {
                "printable_planner": "Printable planners (daily / weekly / monthly)",
                "svg_design": "SVG vector designs (quotes, monograms, patterns)",
                "spreadsheet_template": "Excel spreadsheet templates",
                "social_media_template": "Social-media image templates",
                "wall_art": "Printable wall art (typographic, geometric, abstract)",
                "resume_template": "Resume / CV PDF templates",
                "checklist": "Printable checklists",
                "digital_sticker": "Digital sticker sheets",
            }
        }

    def _ensure_output_dir(self, concept: ProductConcept) -> Path:
        """Create and return the slug-specific output directory."""
        out = self.output_dir / concept.slug
        out.mkdir(parents=True, exist_ok=True)
        return out

    def _get_colors(self, concept: ProductConcept) -> dict:
        """Return the colour palette for the concept."""
        if concept.colors:
            return concept.colors
        return STYLE_PALETTES.get(concept.style, STYLE_PALETTES["minimalist"])

    def _save_metadata(
        self, concept: ProductConcept, output_path: Path, files: list[str]
    ) -> Path:
        """Write a metadata.json beside the generated files."""
        meta = {
            "name": concept.name,
            "slug": concept.slug,
            "category": concept.category,
            "description": concept.description,
            "features": concept.features,
            "style": concept.style,
            "colors": self._get_colors(concept),
            "dimensions": concept.dimensions,
            "pages": concept.pages,
            "target_audience": concept.target_audience,
            "files": files,
            "created_at": datetime.now().isoformat(),
        }
        meta_path = output_path / "metadata.json"
        with open(meta_path, "w") as fh:
            json.dump(meta, fh, indent=2)
        logger.info("Metadata saved → %s", meta_path)
        return meta_path

    # ------------------------------------------------------------------
    # 1. Printable Planner
    # ------------------------------------------------------------------

    def create_printable_planner(self, concept: ProductConcept) -> Path:
        """Generate a multi-page printable planner PDF."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        pr, sr, ar = _hex_to_rgb(colors["primary"]), _hex_to_rgb(colors["secondary"]), _hex_to_rgb(colors["accent"])
        txt_rgb = _hex_to_rgb(colors.get("text", "#2D3436"))

        features_lower = [f.lower() for f in concept.features]
        if "daily" in features_lower:
            layout = "daily"
        elif "monthly" in features_lower:
            layout = "monthly"
        else:
            layout = "weekly"

        w_mm = concept.dimensions.get("width", 612) * 0.3528
        h_mm = concept.dimensions.get("height", 792) * 0.3528
        if w_mm < 50 or h_mm < 50:
            w_mm, h_mm = 215.9, 279.4

        pdf = FPDF(unit="mm", format=(w_mm, h_mm))
        pdf.set_auto_page_break(auto=False)

        # ---- Cover page ----
        pdf.add_page()
        pdf.set_fill_color(*pr)
        pdf.rect(0, 0, w_mm, 80, "F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 28)
        pdf.set_y(22)
        pdf.cell(w_mm, 14, concept.name, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 14)
        pdf.set_text_color(*sr)
        pdf.cell(w_mm, 10, concept.description[:60], align="C", new_x="LMARGIN", new_y="NEXT")
        # decorative accent line
        pdf.set_draw_color(*ar)
        pdf.set_line_width(0.8)
        y_line = 85
        pdf.line(w_mm * 0.2, y_line, w_mm * 0.8, y_line)
        # year
        pdf.set_text_color(*txt_rgb)
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_y(h_mm - 40)
        pdf.cell(w_mm, 12, str(datetime.now().year), align="C")

        # ---- Interior pages ----
        months = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ]
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

        for page_idx in range(concept.pages):
            pdf.add_page()
            # top accent line
            pdf.set_draw_color(*ar)
            pdf.set_line_width(0.4)
            pdf.line(10, 5, w_mm - 10, 5)
            margin = 10
            usable_w = w_mm - 2 * margin

            if layout == "daily":
                pdf.set_font("Helvetica", "B", 16)
                pdf.set_text_color(*pr)
                pdf.set_xy(margin, 10)
                pdf.cell(usable_w, 10, f"Day {page_idx + 1}  —  Date: _______________", new_x="LMARGIN", new_y="NEXT")
                pdf.set_draw_color(*sr)
                pdf.set_line_width(0.2)
                # time slots 6:00 – 21:00
                slot_y = 25
                slot_h = 10
                right_box_w = 60
                time_col_w = 20
                line_col_w = usable_w - right_box_w - time_col_w - 5
                for hour in range(6, 22):
                    pdf.set_font("Helvetica", "", 9)
                    pdf.set_text_color(*sr)
                    pdf.set_xy(margin, slot_y)
                    pdf.cell(time_col_w, slot_h, f"{hour:02d}:00")
                    pdf.set_draw_color(200, 200, 200)
                    pdf.line(margin + time_col_w, slot_y + slot_h, margin + time_col_w + line_col_w, slot_y + slot_h)
                    slot_y += slot_h
                # Priorities box
                box_x = w_mm - margin - right_box_w
                box_y = 25
                box_h = 100
                pdf.set_draw_color(*pr)
                pdf.set_line_width(0.3)
                pdf.rect(box_x, box_y, right_box_w, box_h)
                pdf.set_font("Helvetica", "B", 11)
                pdf.set_text_color(*pr)
                pdf.set_xy(box_x + 2, box_y + 2)
                pdf.cell(right_box_w - 4, 7, "Priorities")
                for i in range(1, 6):
                    py = box_y + 12 + i * 14
                    pdf.set_xy(box_x + 4, py)
                    pdf.set_font("Helvetica", "", 9)
                    pdf.set_text_color(*sr)
                    pdf.cell(5, 5, str(i) + ".")
                    pdf.line(box_x + 12, py + 5, box_x + right_box_w - 4, py + 5)
                # Notes section
                notes_y = h_mm - 50
                pdf.set_draw_color(*pr)
                pdf.rect(margin, notes_y, usable_w, 40)
                pdf.set_font("Helvetica", "B", 11)
                pdf.set_text_color(*pr)
                pdf.set_xy(margin + 2, notes_y + 2)
                pdf.cell(40, 6, "Notes")
                for nl in range(4):
                    ly = notes_y + 12 + nl * 8
                    pdf.set_draw_color(220, 220, 220)
                    pdf.line(margin + 4, ly, margin + usable_w - 4, ly)
                # Water tracker
                water_y = notes_y - 15
                pdf.set_font("Helvetica", "B", 9)
                pdf.set_text_color(*ar)
                pdf.set_xy(margin, water_y)
                pdf.cell(30, 5, "Water Tracker:")
                for ci in range(8):
                    cx = margin + 35 + ci * 12
                    pdf.set_draw_color(*ar)
                    pdf.set_line_width(0.3)
                    pdf.ellipse(cx, water_y, 8, 8)

            elif layout == "weekly":
                pdf.set_font("Helvetica", "B", 16)
                pdf.set_text_color(*pr)
                pdf.set_xy(margin, 10)
                pdf.cell(usable_w, 10, f"Week {page_idx + 1}", new_x="LMARGIN", new_y="NEXT")
                goals_w = 40
                grid_w = usable_w - goals_w - 3
                col_w = grid_w / 7
                header_y = 25
                grid_top = header_y + 8
                grid_bottom = h_mm - 45
                row_h = 15
                # day headers
                for di, day in enumerate(days):
                    cx = margin + di * col_w
                    pdf.set_fill_color(*pr)
                    pdf.set_text_color(255, 255, 255)
                    pdf.set_font("Helvetica", "B", 9)
                    pdf.set_xy(cx, header_y)
                    pdf.cell(col_w, 8, day, border=1, fill=True, align="C")
                # grid lines
                pdf.set_draw_color(210, 210, 210)
                pdf.set_line_width(0.15)
                y_pos = grid_top
                while y_pos <= grid_bottom:
                    pdf.line(margin, y_pos, margin + grid_w, y_pos)
                    y_pos += row_h
                for di in range(8):
                    x_pos = margin + di * col_w
                    pdf.line(x_pos, grid_top, x_pos, min(y_pos, grid_bottom))
                # Goals sidebar
                gs_x = margin + grid_w + 3
                pdf.set_draw_color(*pr)
                pdf.set_line_width(0.3)
                pdf.rect(gs_x, header_y, goals_w, grid_bottom - header_y)
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(*pr)
                pdf.set_xy(gs_x + 2, header_y + 2)
                pdf.cell(goals_w - 4, 7, "Goals")
                for gi in range(6):
                    gy = header_y + 14 + gi * 14
                    pdf.set_draw_color(200, 200, 200)
                    pdf.line(gs_x + 4, gy, gs_x + goals_w - 4, gy)
                # Habit tracker row at bottom
                ht_y = h_mm - 35
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(*ar)
                pdf.set_xy(margin, ht_y)
                pdf.cell(30, 6, "Habit Tracker")
                for hi in range(7):
                    bx = margin + 35 + hi * 22
                    pdf.set_draw_color(*ar)
                    pdf.set_line_width(0.3)
                    pdf.rect(bx, ht_y, 6, 6)
                    pdf.set_font("Helvetica", "", 7)
                    pdf.set_text_color(*sr)
                    pdf.set_xy(bx + 7, ht_y)
                    pdf.cell(14, 6, days[hi])

            else:  # monthly
                month_name = months[page_idx % 12]
                pdf.set_font("Helvetica", "B", 18)
                pdf.set_text_color(*pr)
                pdf.set_xy(margin, 10)
                pdf.cell(usable_w * 0.7, 10, month_name, new_x="LMARGIN", new_y="NEXT")
                # Calendar grid
                cal_top = 28
                goals_w = 45
                cal_w = usable_w - goals_w - 4
                col_w = cal_w / 7
                row_h = 18
                # day headers
                for di, day in enumerate(days):
                    cx = margin + di * col_w
                    pdf.set_fill_color(*pr)
                    pdf.set_text_color(255, 255, 255)
                    pdf.set_font("Helvetica", "B", 8)
                    pdf.set_xy(cx, cal_top)
                    pdf.cell(col_w, 7, day, border=1, fill=True, align="C")
                # 6 rows × 7 cols
                pdf.set_draw_color(200, 200, 200)
                pdf.set_line_width(0.15)
                day_num = 1
                for row in range(6):
                    for col in range(7):
                        cx = margin + col * col_w
                        cy = cal_top + 7 + row * row_h
                        pdf.rect(cx, cy, col_w, row_h)
                        if day_num <= 31:
                            pdf.set_font("Helvetica", "", 8)
                            pdf.set_text_color(*txt_rgb)
                            pdf.set_xy(cx + 1, cy + 1)
                            pdf.cell(col_w - 2, 5, str(day_num))
                            day_num += 1
                # Notes section
                notes_y = cal_top + 7 + 6 * row_h + 5
                pdf.set_draw_color(*pr)
                pdf.set_line_width(0.3)
                pdf.rect(margin, notes_y, cal_w, 35)
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(*pr)
                pdf.set_xy(margin + 2, notes_y + 2)
                pdf.cell(30, 6, "Notes")
                for nl in range(3):
                    ly = notes_y + 11 + nl * 8
                    pdf.set_draw_color(220, 220, 220)
                    pdf.line(margin + 4, ly, margin + cal_w - 4, ly)
                # Monthly Goals sidebar
                gs_x = margin + cal_w + 4
                gs_top = cal_top
                gs_h = notes_y + 35 - gs_top
                pdf.set_draw_color(*pr)
                pdf.rect(gs_x, gs_top, goals_w, gs_h)
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(*pr)
                pdf.set_xy(gs_x + 2, gs_top + 2)
                pdf.cell(goals_w - 4, 7, "Monthly Goals")
                for gi in range(8):
                    gy = gs_top + 14 + gi * 14
                    if gy < gs_top + gs_h - 5:
                        pdf.set_draw_color(200, 200, 200)
                        pdf.line(gs_x + 4, gy, gs_x + goals_w - 4, gy)

            # page number
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(160, 160, 160)
            pdf.set_xy(0, h_mm - 10)
            pdf.cell(w_mm, 5, str(page_idx + 2), align="C")

        pdf_path = out_dir / f"{concept.slug}.pdf"
        pdf.output(str(pdf_path))
        logger.info("Planner PDF (%s layout, %d pages) → %s", layout, concept.pages, pdf_path)
        return pdf_path

    # ------------------------------------------------------------------
    # 2. SVG Design
    # ------------------------------------------------------------------

    def create_svg_design(self, concept: ProductConcept) -> Path:
        """Generate a decorative SVG design."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        w = concept.dimensions.get("width", 612)
        h = concept.dimensions.get("height", 792)
        svg_path = out_dir / f"{concept.slug}.svg"

        dwg = svgwrite.Drawing(str(svg_path), size=(f"{w}px", f"{h}px"), viewBox=f"0 0 {w} {h}")

        features_lower = " ".join(concept.features + [concept.description]).lower()

        if "quote" in features_lower:
            # ornamental border
            dwg.add(dwg.rect(insert=(20, 20), size=(w - 40, h - 40), rx=12, ry=12,
                             fill="none", stroke=colors["primary"], stroke_width=3))
            dwg.add(dwg.rect(insert=(30, 30), size=(w - 60, h - 60), rx=8, ry=8,
                             fill="none", stroke=colors["accent"], stroke_width=1))
            # decorative corner flourishes
            for cx, cy, sx, sy in [(40, 40, 1, 1), (w - 40, 40, -1, 1),
                                    (40, h - 40, 1, -1), (w - 40, h - 40, -1, -1)]:
                path_d = f"M {cx},{cy} c {sx * 30},0 0,{sy * 30} {sx * 30},{sy * 30}"
                dwg.add(dwg.path(d=path_d, fill="none", stroke=colors["accent"], stroke_width=2))
            # quote text
            lines = concept.name.split()
            mid = len(lines) // 2 or 1
            line1 = " ".join(lines[:mid])
            line2 = " ".join(lines[mid:])
            text_y = h // 2 - 20
            dwg.add(dwg.text(line1, insert=(w // 2, text_y),
                             text_anchor="middle", font_size="36px", font_family="Georgia",
                             fill=colors["primary"], font_weight="bold"))
            if line2:
                dwg.add(dwg.text(line2, insert=(w // 2, text_y + 48),
                                 text_anchor="middle", font_size="36px", font_family="Georgia",
                                 fill=colors["primary"], font_weight="bold"))
            # decorative lines above/below
            dwg.add(dwg.line(start=(w * 0.25, text_y - 30), end=(w * 0.75, text_y - 30),
                             stroke=colors["accent"], stroke_width=2))
            dwg.add(dwg.line(start=(w * 0.25, text_y + 70), end=(w * 0.75, text_y + 70),
                             stroke=colors["accent"], stroke_width=2))

        elif "monogram" in features_lower:
            # decorative frame
            cx, cy = w // 2, h // 2
            for r, color, sw in [(min(w, h) * 0.35, colors["primary"], 4),
                                  (min(w, h) * 0.32, colors["accent"], 2),
                                  (min(w, h) * 0.38, colors["secondary"], 1)]:
                dwg.add(dwg.circle(center=(cx, cy), r=r, fill="none", stroke=color, stroke_width=sw))
            letter = concept.name[0].upper() if concept.name else "A"
            dwg.add(dwg.text(letter, insert=(cx, cy + 40),
                             text_anchor="middle", font_size="120px",
                             font_family="Georgia", fill=colors["primary"], font_weight="bold"))
            # ornamental dots around circle
            radius = min(w, h) * 0.38 + 15
            for i in range(24):
                angle = math.radians(i * 15)
                dx = cx + radius * math.cos(angle)
                dy = cy + radius * math.sin(angle)
                dwg.add(dwg.circle(center=(dx, dy), r=3, fill=colors["accent"]))

        else:
            # geometric / floral pattern
            bg_rgb = colors.get("background", "#FFFFFF")
            dwg.add(dwg.rect(insert=(0, 0), size=(w, h), fill=bg_rgb))
            palette = [colors["primary"], colors["secondary"], colors["accent"]]
            step = 60
            idx = 0
            for row in range(0, h, step):
                for col in range(0, w, step):
                    c = palette[idx % len(palette)]
                    idx += 1
                    cx_p = col + step // 2
                    cy_p = row + step // 2
                    if (row // step + col // step) % 3 == 0:
                        dwg.add(dwg.circle(center=(cx_p, cy_p), r=step * 0.3,
                                           fill="none", stroke=c, stroke_width=2))
                    elif (row // step + col // step) % 3 == 1:
                        half = step * 0.25
                        dwg.add(dwg.rect(insert=(cx_p - half, cy_p - half),
                                         size=(half * 2, half * 2),
                                         fill="none", stroke=c, stroke_width=2,
                                         transform=f"rotate(45,{cx_p},{cy_p})"))
                    else:
                        r = step * 0.28
                        points = []
                        for pi in range(6):
                            angle = math.radians(60 * pi - 30)
                            points.append((cx_p + r * math.cos(angle), cy_p + r * math.sin(angle)))
                        dwg.add(dwg.polygon(points, fill="none", stroke=c, stroke_width=2))

        dwg.save()
        logger.info("SVG design → %s", svg_path)
        return svg_path

    # ------------------------------------------------------------------
    # 3. Spreadsheet
    # ------------------------------------------------------------------

    def create_spreadsheet(self, concept: ProductConcept) -> Path:
        """Generate an Excel spreadsheet template with three sheets."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        xlsx_path = out_dir / f"{concept.slug}.xlsx"

        pr_hex = colors["primary"].lstrip("#")
        sr_hex = colors["secondary"].lstrip("#")
        ac_hex = colors["accent"].lstrip("#")

        wb = openpyxl.Workbook()

        # ---- Overview sheet ----
        ws_ov = wb.active
        ws_ov.title = "Overview"
        ws_ov.merge_cells("A1:F1")
        title_cell = ws_ov["A1"]
        title_cell.value = concept.name
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=pr_hex, end_color=pr_hex, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_ov.row_dimensions[1].height = 36

        instructions = [
            f"Product: {concept.name}",
            f"Category: {concept.category}",
            f"Description: {concept.description}",
            "",
            "HOW TO USE THIS TEMPLATE:",
            "1. Navigate to the 'Data Entry' sheet to input your data.",
            "2. Fill in all required columns (highlighted in the header row).",
            "3. The 'Dashboard' sheet updates automatically with summary stats.",
            "",
            "Tip: Use Ctrl+D to copy a cell value downward.",
        ]
        for i, line in enumerate(instructions, start=3):
            ws_ov.cell(row=i, column=1, value=line)
        # Legend
        ws_ov.cell(row=14, column=1, value="Legend / Colour Key:").font = Font(bold=True)
        legend = [("Primary", pr_hex), ("Secondary", sr_hex), ("Accent", ac_hex)]
        for li, (label, hex_c) in enumerate(legend, start=15):
            ws_ov.cell(row=li, column=1, value=label)
            ws_ov.cell(row=li, column=2).fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
        ws_ov.column_dimensions["A"].width = 45

        # ---- Data Entry sheet ----
        ws_de = wb.create_sheet("Data Entry")
        headers = ["ID", "Date", "Category", "Item", "Amount", "Notes"]
        header_fill = PatternFill(start_color=pr_hex, end_color=pr_hex, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        for ci, hdr in enumerate(headers, start=1):
            cell = ws_de.cell(row=1, column=ci, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        ws_de.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws_de.freeze_panes = "A2"
        col_widths = {"A": 8, "B": 14, "C": 18, "D": 25, "E": 14, "F": 30}
        for col_letter, width in col_widths.items():
            ws_de.column_dimensions[col_letter].width = width
        # sample data
        sample_rows = [
            [1, "2025-01-01", "Revenue", "Product A", 149.99, "Etsy sale"],
            [2, "2025-01-02", "Expense", "Supplies", -25.00, "Packaging materials"],
            [3, "2025-01-03", "Revenue", "Product B", 79.50, "Direct sale"],
            [4, "2025-01-04", "Revenue", "Product A", 149.99, "Etsy sale"],
            [5, "2025-01-05", "Expense", "Marketing", -50.00, "Facebook ads"],
        ]
        for ri, row_data in enumerate(sample_rows, start=2):
            for ci, val in enumerate(row_data, start=1):
                cell = ws_de.cell(row=ri, column=ci, value=val)
                cell.border = thin_border

        # ---- Dashboard sheet ----
        ws_db = wb.create_sheet("Dashboard")
        ws_db.merge_cells("A1:D1")
        dash_title = ws_db["A1"]
        dash_title.value = "Dashboard — Summary"
        dash_title.font = Font(bold=True, size=14, color="FFFFFF")
        dash_title.fill = PatternFill(start_color=pr_hex, end_color=pr_hex, fill_type="solid")
        dash_title.alignment = Alignment(horizontal="center")
        ws_db.row_dimensions[1].height = 30

        summary_items = [
            ("Total Records", "=COUNTA('Data Entry'!A2:A1000)"),
            ("Sum of Amounts", "=SUM('Data Entry'!E2:E1000)"),
            ("Average Amount", "=AVERAGE('Data Entry'!E2:E1000)"),
            ("Max Amount", "=MAX('Data Entry'!E2:E1000)"),
            ("Min Amount", "=MIN('Data Entry'!E2:E1000)"),
        ]
        accent_fill = PatternFill(start_color=ac_hex, end_color=ac_hex, fill_type="solid")
        for si, (label, formula) in enumerate(summary_items, start=3):
            lbl_cell = ws_db.cell(row=si, column=1, value=label)
            lbl_cell.font = Font(bold=True)
            lbl_cell.border = thin_border
            val_cell = ws_db.cell(row=si, column=2, value=formula)
            val_cell.border = thin_border
            val_cell.fill = accent_fill
            val_cell.number_format = "#,##0.00"
        ws_db.column_dimensions["A"].width = 22
        ws_db.column_dimensions["B"].width = 18

        wb.save(str(xlsx_path))
        logger.info("Spreadsheet → %s", xlsx_path)
        return xlsx_path

    # ------------------------------------------------------------------
    # 4. Social-media Template
    # ------------------------------------------------------------------

    def create_social_template(self, concept: ProductConcept) -> Path:
        """Generate a social-media image template using Pillow."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        pr, sr, ar = _hex_to_rgb(colors["primary"]), _hex_to_rgb(colors["secondary"]), _hex_to_rgb(colors["accent"])

        features_lower = " ".join(concept.features).lower()
        if "instagram_story" in features_lower or "story" in features_lower:
            size = (1080, 1920)
        elif "pinterest" in features_lower or "pin" in features_lower:
            size = (1000, 1500)
        else:
            size = (1080, 1080)

        img = Image.new("RGB", size, pr)
        draw = ImageDraw.Draw(img)

        # vertical gradient
        for y in range(size[1]):
            ratio = y / size[1]
            r = int(pr[0] * (1 - ratio) + sr[0] * ratio)
            g = int(pr[1] * (1 - ratio) + sr[1] * ratio)
            b = int(pr[2] * (1 - ratio) + sr[2] * ratio)
            draw.line([(0, y), (size[0], y)], fill=(r, g, b))

        # semi-transparent text area in centre
        overlay = Image.new("RGBA", size, (0, 0, 0, 0))
        ov_draw = ImageDraw.Draw(overlay)
        pad = int(size[0] * 0.1)
        box_y1 = int(size[1] * 0.25)
        box_y2 = int(size[1] * 0.65)
        ov_draw.rounded_rectangle(
            [pad, box_y1, size[0] - pad, box_y2],
            radius=20, fill=(255, 255, 255, 180),
        )
        img = img.convert("RGBA")
        img = Image.alpha_composite(img, overlay)
        draw = ImageDraw.Draw(img)

        # text labels
        try:
            title_font = ImageFont.truetype("Arial.ttf", int(size[0] * 0.06))
            sub_font = ImageFont.truetype("Arial.ttf", int(size[0] * 0.035))
        except (IOError, OSError):
            title_font = ImageFont.load_default()
            sub_font = ImageFont.load_default()

        title_text = "YOUR TEXT HERE"
        sub_text = "SUBTITLE"
        title_bbox = draw.textbbox((0, 0), title_text, font=title_font)
        tw = title_bbox[2] - title_bbox[0]
        tx = (size[0] - tw) // 2
        ty = box_y1 + (box_y2 - box_y1) // 2 - 40
        draw.text((tx, ty), title_text, fill=(*pr, 255), font=title_font)

        sub_bbox = draw.textbbox((0, 0), sub_text, font=sub_font)
        sw = sub_bbox[2] - sub_bbox[0]
        sx = (size[0] - sw) // 2
        draw.text((sx, ty + 60), sub_text, fill=(*sr, 255), font=sub_font)

        # decorative corners
        corner_len = int(size[0] * 0.08)
        corner_w = 4
        ac_rgba = (*ar, 255)
        corners = [
            (pad - 10, box_y1 - 10),
            (size[0] - pad + 10 - corner_len, box_y1 - 10),
            (pad - 10, box_y2 + 10 - corner_len),
            (size[0] - pad + 10 - corner_len, box_y2 + 10 - corner_len),
        ]
        for cx, cy in corners:
            draw.line([(cx, cy), (cx + corner_len, cy)], fill=ac_rgba, width=corner_w)
            draw.line([(cx, cy), (cx, cy + corner_len)], fill=ac_rgba, width=corner_w)

        # bottom branding bar
        bar_h = int(size[1] * 0.06)
        draw.rectangle([0, size[1] - bar_h, size[0], size[1]], fill=(*pr, 220))
        try:
            brand_font = ImageFont.truetype("Arial.ttf", int(bar_h * 0.4))
        except (IOError, OSError):
            brand_font = ImageFont.load_default()
        brand_text = f"@{concept.slug}"
        bb = draw.textbbox((0, 0), brand_text, font=brand_font)
        bw = bb[2] - bb[0]
        draw.text(((size[0] - bw) // 2, size[1] - bar_h + int(bar_h * 0.25)),
                  brand_text, fill=(255, 255, 255, 255), font=brand_font)

        png_path = out_dir / f"{concept.slug}.png"
        img.convert("RGB").save(str(png_path), "PNG")
        logger.info("Social template (%dx%d) → %s", size[0], size[1], png_path)
        return png_path

    # ------------------------------------------------------------------
    # 5. Wall Art
    # ------------------------------------------------------------------

    def create_wall_art(self, concept: ProductConcept) -> Path:
        """Generate printable wall art at 300 DPI."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        pr, sr, ar = _hex_to_rgb(colors["primary"]), _hex_to_rgb(colors["secondary"]), _hex_to_rgb(colors["accent"])
        bg = _hex_to_rgb(colors.get("background", "#FFFFFF"))

        w = concept.dimensions.get("width", 2400)
        h = concept.dimensions.get("height", 3000)
        if w < 500:
            w, h = 2400, 3000

        features_lower = " ".join(concept.features + [concept.style]).lower()

        img = Image.new("RGB", (w, h), bg)
        draw = ImageDraw.Draw(img)

        if "geometric" in features_lower:
            # grid of circles/squares
            step_x = w // 10
            step_y = h // 12
            palette = [pr, sr, ar]
            idx = 0
            for row in range(12):
                for col in range(10):
                    cx = col * step_x + step_x // 2
                    cy = row * step_y + step_y // 2
                    c = palette[idx % len(palette)]
                    idx += 1
                    radius = min(step_x, step_y) // 3
                    if (row + col) % 2 == 0:
                        draw.ellipse(
                            [cx - radius, cy - radius, cx + radius, cy + radius],
                            fill=c,
                        )
                    else:
                        half = radius
                        draw.rectangle(
                            [cx - half, cy - half, cx + half, cy + half],
                            fill=c,
                        )

        elif "abstract" in features_lower:
            # gradient background
            for y in range(h):
                ratio = y / h
                r = int(pr[0] * (1 - ratio) + sr[0] * ratio)
                g = int(pr[1] * (1 - ratio) + sr[1] * ratio)
                b = int(pr[2] * (1 - ratio) + sr[2] * ratio)
                draw.line([(0, y), (w, y)], fill=(r, g, b))
            # overlay shapes
            import random
            rng = random.Random(42)  # deterministic
            for _ in range(30):
                shape_x = rng.randint(0, w)
                shape_y = rng.randint(0, h)
                shape_r = rng.randint(40, 200)
                alpha_val = rng.randint(40, 120)
                shape_c = list(palette[rng.randint(0, 2)] if 'palette' in dir() else ar) + [alpha_val]
                # draw on RGBA overlay
                pass
            # simpler approach: draw outlined circles and rectangles
            for i in range(25):
                cx = int(w * ((i * 7 + 13) % 100) / 100)
                cy = int(h * ((i * 11 + 29) % 100) / 100)
                radius = 50 + (i * 17) % 150
                c = [pr, sr, ar][i % 3]
                draw.ellipse(
                    [cx - radius, cy - radius, cx + radius, cy + radius],
                    outline=(*c, ), width=3,
                )
            for i in range(15):
                rx = int(w * ((i * 13 + 7) % 100) / 100)
                ry = int(h * ((i * 19 + 3) % 100) / 100)
                rw = 80 + (i * 23) % 120
                rh = 80 + (i * 31) % 120
                c = [pr, sr, ar][i % 3]
                draw.rectangle([rx, ry, rx + rw, ry + rh], outline=c, width=2)

        else:
            # typographic wall art
            text = concept.name
            try:
                font_size = int(min(w, h) * 0.08)
                font = ImageFont.truetype("Arial Bold.ttf", font_size)
            except (IOError, OSError):
                try:
                    font_size = int(min(w, h) * 0.08)
                    font = ImageFont.truetype("Arial.ttf", font_size)
                except (IOError, OSError):
                    font = ImageFont.load_default()

            # split text into lines that fit
            words = text.split()
            lines: list[str] = []
            current = ""
            for word in words:
                test = f"{current} {word}".strip()
                bbox = draw.textbbox((0, 0), test, font=font)
                if bbox[2] - bbox[0] > w * 0.8 and current:
                    lines.append(current)
                    current = word
                else:
                    current = test
            if current:
                lines.append(current)
            if not lines:
                lines = [text]

            line_height = int(font_size * 1.4)
            total_h = line_height * len(lines)
            start_y = (h - total_h) // 2

            # decorative line above
            draw.line([(w * 0.2, start_y - 40), (w * 0.8, start_y - 40)], fill=ar, width=4)

            for li, line in enumerate(lines):
                bbox = draw.textbbox((0, 0), line, font=font)
                lw = bbox[2] - bbox[0]
                lx = (w - lw) // 2
                ly = start_y + li * line_height
                draw.text((lx, ly), line, fill=pr, font=font)

            # decorative line below
            draw.line([(w * 0.2, start_y + total_h + 20), (w * 0.8, start_y + total_h + 20)], fill=ar, width=4)

            # small decorative elements
            dot_r = 8
            for dx in range(5):
                x = int(w * 0.3 + dx * w * 0.1)
                draw.ellipse([x - dot_r, start_y - 70 - dot_r, x + dot_r, start_y - 70 + dot_r], fill=sr)
                draw.ellipse([x - dot_r, start_y + total_h + 50 - dot_r, x + dot_r, start_y + total_h + 50 + dot_r], fill=sr)

        img.info["dpi"] = (300, 300)
        png_path = out_dir / f"{concept.slug}.png"
        img.save(str(png_path), "PNG", dpi=(300, 300))
        logger.info("Wall art (%dx%d @300dpi) → %s", w, h, png_path)
        return png_path

    # ------------------------------------------------------------------
    # 6. Resume Template
    # ------------------------------------------------------------------

    def create_resume_template(self, concept: ProductConcept) -> Path:
        """Generate a two-column resume PDF template."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        pr, sr, ar = _hex_to_rgb(colors["primary"]), _hex_to_rgb(colors["secondary"]), _hex_to_rgb(colors["accent"])
        txt_rgb = _hex_to_rgb(colors.get("text", "#2D3436"))

        html_content = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Helvetica Neue', Arial, sans-serif; font-size: 11pt; color: {colors.get('text','#2D3436')}; }}
.page {{ display: flex; width: 210mm; min-height: 297mm; }}
.sidebar {{ width: 30%; background: {colors['primary']}; color: #fff; padding: 30px 18px; }}
.sidebar h1 {{ font-size: 22pt; margin-bottom: 5px; }}
.sidebar h3 {{ font-size: 10pt; color: {colors['accent']}; margin: 18px 0 8px; text-transform: uppercase; letter-spacing: 1px; }}
.sidebar p, .sidebar li {{ font-size: 9pt; line-height: 1.6; }}
.sidebar ul {{ list-style: none; }}
.sidebar ul li::before {{ content: "▸ "; color: {colors['accent']}; }}
.main {{ width: 70%; padding: 30px 24px; }}
.main h2 {{ font-size: 13pt; color: {colors['primary']}; border-bottom: 2px solid {colors['accent']}; padding-bottom: 4px; margin: 20px 0 10px; }}
.main h2:first-child {{ margin-top: 0; }}
.entry {{ margin-bottom: 14px; }}
.entry .title {{ font-weight: bold; }}
.entry .meta {{ font-size: 9pt; color: {colors['secondary']}; margin-bottom: 4px; }}
.entry ul {{ margin-left: 16px; font-size: 10pt; line-height: 1.6; }}
</style></head><body><div class="page">
<div class="sidebar">
<h1>[Your Name]</h1>
<p>[Title / Headline]</p>
<h3>Contact</h3>
<p>📧 email@example.com<br>📱 (555) 123-4567<br>📍 City, State<br>🔗 linkedin.com/in/yourname</p>
<h3>Skills</h3>
<ul><li>Skill Category A</li><li>Skill Category B</li><li>Skill Category C</li><li>Skill Category D</li><li>Skill Category E</li></ul>
<h3>Languages</h3>
<ul><li>English — Native</li><li>Spanish — Professional</li></ul>
<h3>Certifications</h3>
<ul><li>Certification One</li><li>Certification Two</li></ul>
</div>
<div class="main">
<h2>Summary</h2>
<p>Results-driven professional with 5+ years of experience in [industry]. Proven track record of delivering high-impact projects, leading cross-functional teams, and driving measurable business outcomes.</p>
<h2>Experience</h2>
<div class="entry"><span class="title">Senior Role Title</span><div class="meta">Company Name  |  Jan 2022 – Present</div>
<ul><li>Led initiative that increased metric by 35% within first quarter</li><li>Managed team of 8 across multiple concurrent projects</li><li>Designed and implemented scalable process improvements</li></ul></div>
<div class="entry"><span class="title">Mid-Level Role Title</span><div class="meta">Previous Company  |  Jun 2019 – Dec 2021</div>
<ul><li>Delivered project ahead of schedule, saving $50K in costs</li><li>Collaborated with stakeholders to define requirements and roadmap</li><li>Mentored 3 junior team members to promotion readiness</li></ul></div>
<h2>Education</h2>
<div class="entry"><span class="title">Bachelor of Science in [Field]</span><div class="meta">University Name  |  Graduated 2019</div>
<ul><li>Dean's List — 4 semesters</li><li>Relevant coursework: Topic A, Topic B, Topic C</li></ul></div>
<h2>Additional Skills</h2>
<p>Tool A • Tool B • Tool C • Framework X • Platform Y • Methodology Z</p>
</div></div></body></html>"""

        # Try WeasyPrint, fall back to FPDF
        pdf_path = out_dir / f"{concept.slug}.pdf"
        try:
            import weasyprint  # type: ignore
            weasyprint.HTML(string=html_content).write_pdf(str(pdf_path))
            logger.info("Resume (WeasyPrint) → %s", pdf_path)
            return pdf_path
        except (ImportError, Exception) as exc:
            logger.debug("WeasyPrint unavailable (%s), falling back to FPDF", exc)

        # FPDF fallback
        pdf = FPDF(unit="mm", format="A4")
        pdf.set_auto_page_break(auto=False)
        pdf.add_page()
        page_w, page_h = 210, 297
        sidebar_w = page_w * 0.30

        # Sidebar background
        pdf.set_fill_color(*pr)
        pdf.rect(0, 0, sidebar_w, page_h, "F")

        # Sidebar content
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_xy(8, 15)
        pdf.cell(sidebar_w - 16, 10, "[Your Name]")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_xy(8, 28)
        pdf.cell(sidebar_w - 16, 6, "[Title / Headline]")

        sections = [
            ("CONTACT", [
                "email@example.com",
                "(555) 123-4567",
                "City, State",
                "linkedin.com/in/you",
            ]),
            ("SKILLS", [
                "Skill Category A",
                "Skill Category B",
                "Skill Category C",
                "Skill Category D",
                "Skill Category E",
            ]),
            ("LANGUAGES", [
                "English - Native",
                "Spanish - Professional",
            ]),
            ("CERTIFICATIONS", [
                "Certification One",
                "Certification Two",
            ]),
        ]
        y = 42
        for title, items in sections:
            pdf.set_text_color(*ar)
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_xy(8, y)
            pdf.cell(sidebar_w - 16, 5, title)
            y += 7
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "", 8)
            for item in items:
                pdf.set_xy(10, y)
                pdf.cell(sidebar_w - 20, 5, f"  {item}")
                y += 6
            y += 4

        # Main content
        main_x = sidebar_w + 8
        main_w = page_w - sidebar_w - 16
        y = 15

        def section_header(title: str, yy: float) -> float:
            pdf.set_draw_color(*ar)
            pdf.set_text_color(*pr)
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_xy(main_x, yy)
            pdf.cell(main_w, 7, title)
            yy += 8
            pdf.set_line_width(0.5)
            pdf.line(main_x, yy, main_x + main_w, yy)
            return yy + 4

        y = section_header("Summary", y)
        pdf.set_text_color(*txt_rgb)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_xy(main_x, y)
        pdf.multi_cell(main_w, 5,
                        "Results-driven professional with 5+ years of experience. "
                        "Proven track record of delivering high-impact projects, "
                        "leading cross-functional teams, and driving measurable outcomes.")
        y = pdf.get_y() + 4

        y = section_header("Experience", y)
        experience = [
            {
                "title": "Senior Role Title",
                "meta": "Company Name  |  Jan 2022 - Present",
                "bullets": [
                    "Led initiative that increased metric by 35%",
                    "Managed team of 8 across concurrent projects",
                    "Designed scalable process improvements",
                ],
            },
            {
                "title": "Mid-Level Role Title",
                "meta": "Previous Company  |  Jun 2019 - Dec 2021",
                "bullets": [
                    "Delivered project ahead of schedule, saving $50K",
                    "Collaborated with stakeholders on requirements",
                    "Mentored 3 junior team members to promotion",
                ],
            },
        ]
        for entry in experience:
            pdf.set_text_color(*txt_rgb)
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_xy(main_x, y)
            pdf.cell(main_w, 6, entry["title"])
            y += 6
            pdf.set_text_color(*sr)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_xy(main_x, y)
            pdf.cell(main_w, 5, entry["meta"])
            y += 6
            pdf.set_text_color(*txt_rgb)
            pdf.set_font("Helvetica", "", 8)
            for bullet in entry["bullets"]:
                pdf.set_xy(main_x + 4, y)
                pdf.cell(main_w - 4, 5, f"  {bullet}")
                y += 5
            y += 5

        y = section_header("Education", y)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*txt_rgb)
        pdf.set_xy(main_x, y)
        pdf.cell(main_w, 6, "Bachelor of Science in [Field]")
        y += 6
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(*sr)
        pdf.set_xy(main_x, y)
        pdf.cell(main_w, 5, "University Name  |  Graduated 2019")
        y += 7
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*txt_rgb)
        for bullet in ["Dean's List - 4 semesters", "Relevant coursework: Topic A, Topic B, Topic C"]:
            pdf.set_xy(main_x + 4, y)
            pdf.cell(main_w - 4, 5, f"  {bullet}")
            y += 5
        y += 5

        y = section_header("Additional Skills", y)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*txt_rgb)
        pdf.set_xy(main_x, y)
        pdf.cell(main_w, 6, "Tool A  |  Tool B  |  Framework X  |  Platform Y")

        pdf.output(str(pdf_path))
        logger.info("Resume (FPDF fallback) → %s", pdf_path)
        return pdf_path

    # ------------------------------------------------------------------
    # 7. Checklist
    # ------------------------------------------------------------------

    def create_checklist(self, concept: ProductConcept) -> Path:
        """Generate a printable checklist PDF."""
        out_dir = self._ensure_output_dir(concept)
        colors = self._get_colors(concept)
        pr, sr, ar = _hex_to_rgb(colors["primary"]), _hex_to_rgb(colors["secondary"]), _hex_to_rgb(colors["accent"])
        txt_rgb = _hex_to_rgb(colors.get("text", "#2D3436"))

        pdf = FPDF(unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        w_mm, h_mm = 210, 297
        margin = 10

        pdf.add_page()
        # Title band
        pdf.set_fill_color(*pr)
        pdf.rect(0, 0, w_mm, 25, "F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_xy(margin, 5)
        pdf.cell(w_mm - 2 * margin, 15, concept.name, align="C")

        # Derive sections from features or defaults
        if concept.features:
            sections_raw = concept.features
        else:
            sections_raw = ["Priority Items", "In Progress", "Completed"]

        # Build checklist items per section
        default_items = [
            "Item one — describe the task",
            "Item two — describe the task",
            "Item three — describe the task",
            "Item four — describe the task",
            "Item five — describe the task",
        ]

        y = 32
        page_num = 1

        for sec_idx, section_name in enumerate(sections_raw):
            # section header
            pdf.set_fill_color(*ar)
            pdf.set_draw_color(*ar)
            pdf.rect(margin, y, w_mm - 2 * margin, 8, "F")
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_xy(margin + 3, y + 1)
            pdf.cell(w_mm - 2 * margin - 6, 6, section_name.title())
            y += 12

            # checklist items
            pdf.set_text_color(*txt_rgb)
            pdf.set_font("Helvetica", "", 10)
            for item in default_items:
                if y > h_mm - 25:
                    # page footer
                    pdf.set_font("Helvetica", "", 8)
                    pdf.set_text_color(160, 160, 160)
                    pdf.set_xy(0, h_mm - 12)
                    pdf.cell(w_mm, 5, str(page_num), align="C")
                    page_num += 1
                    pdf.add_page()
                    y = 15
                # checkbox character and item text
                pdf.set_xy(margin + 4, y)
                pdf.set_font("Helvetica", "", 12)
                pdf.set_text_color(*sr)
                pdf.cell(6, 6, chr(9744))  # ☐
                pdf.set_font("Helvetica", "", 10)
                pdf.set_text_color(*txt_rgb)
                pdf.set_xy(margin + 12, y)
                pdf.cell(w_mm - 2 * margin - 16, 6, item)
                # light underline
                pdf.set_draw_color(220, 220, 220)
                pdf.set_line_width(0.15)
                pdf.line(margin + 12, y + 7, w_mm - margin, y + 7)
                y += 8

            y += 6  # gap between sections

        # final page number
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(160, 160, 160)
        pdf.set_xy(0, h_mm - 12)
        pdf.cell(w_mm, 5, str(page_num), align="C")

        pdf_path = out_dir / f"{concept.slug}.pdf"
        pdf.output(str(pdf_path))
        logger.info("Checklist → %s", pdf_path)
        return pdf_path

    # ------------------------------------------------------------------
    # 8. Mockup Generator
    # ------------------------------------------------------------------

    def generate_mockup(self, product_path: Path, style: str = "default") -> Path:
        """Create an Etsy-style mockup image for a product file."""
        product_path = Path(product_path)
        canvas_w, canvas_h = 1500, 1200

        # warm gradient background (wood tones)
        bg_top = (222, 184, 135)   # #DEB887
        bg_bot = (210, 180, 140)   # #D2B48C
        canvas = Image.new("RGB", (canvas_w, canvas_h), bg_top)
        draw = ImageDraw.Draw(canvas)
        for y in range(canvas_h):
            ratio = y / canvas_h
            r = int(bg_top[0] * (1 - ratio) + bg_bot[0] * ratio)
            g = int(bg_top[1] * (1 - ratio) + bg_bot[1] * ratio)
            b = int(bg_top[2] * (1 - ratio) + bg_bot[2] * ratio)
            draw.line([(0, y), (canvas_w, y)], fill=(r, g, b))

        # subtle wood grain texture
        import random
        rng = random.Random(12345)
        for _ in range(4000):
            tx = rng.randint(0, canvas_w - 1)
            ty = rng.randint(0, canvas_h - 1)
            base = canvas.getpixel((tx, ty))
            offset = rng.randint(-8, 8)
            dot_c = tuple(max(0, min(255, c + offset)) for c in base)
            draw.point((tx, ty), fill=dot_c)

        # load product image or create placeholder
        ext = product_path.suffix.lower()
        if ext in (".png", ".jpg", ".jpeg", ".webp"):
            try:
                prod_img = Image.open(product_path).convert("RGBA")
            except Exception:
                prod_img = Image.new("RGBA", (600, 800), (255, 255, 255, 255))
                pd = ImageDraw.Draw(prod_img)
                pd.text((50, 380), product_path.name, fill=(100, 100, 100, 255))
        else:
            # PDF or other: create a white placeholder with filename
            prod_img = Image.new("RGBA", (600, 800), (255, 255, 255, 255))
            pd = ImageDraw.Draw(prod_img)
            try:
                pf = ImageFont.truetype("Arial.ttf", 24)
            except (IOError, OSError):
                pf = ImageFont.load_default()
            pd.text((40, 60), product_path.stem, fill=(60, 60, 60, 255), font=pf)
            pd.text((40, 100), f"[{ext.upper().lstrip('.')} Preview]", fill=(120, 120, 120, 255), font=pf)
            # draw decorative border
            pd.rectangle([10, 10, 590, 790], outline=(200, 200, 200, 255), width=2)

        # Resize to fit 60% of canvas
        max_w = int(canvas_w * 0.6)
        max_h = int(canvas_h * 0.6)
        prod_img.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)
        pw, ph = prod_img.size

        # drop shadow
        shadow_offset = 12
        shadow = Image.new("RGBA", (pw + 30, ph + 30), (0, 0, 0, 0))
        sd = ImageDraw.Draw(shadow)
        sd.rectangle([shadow_offset, shadow_offset, pw + shadow_offset, ph + shadow_offset],
                      fill=(0, 0, 0, 80))
        shadow = shadow.filter(ImageFilter.GaussianBlur(radius=10))

        # position product centred on canvas
        pos_x = (canvas_w - pw) // 2
        pos_y = (canvas_h - ph) // 2 - 30
        canvas.paste(Image.new("RGB", shadow.size, (0, 0, 0)),
                      (pos_x - 5, pos_y - 5),
                      shadow.split()[3])
        canvas.paste(prod_img, (pos_x, pos_y),
                      prod_img.split()[3] if prod_img.mode == "RGBA" else None)

        # branding text at bottom
        try:
            brand_font = ImageFont.truetype("Arial.ttf", 22)
        except (IOError, OSError):
            brand_font = ImageFont.load_default()
        brand_text = "Digital Download  •  Instant Delivery"
        bb = draw.textbbox((0, 0), brand_text, font=brand_font)
        tw = bb[2] - bb[0]
        draw = ImageDraw.Draw(canvas)
        draw.text(((canvas_w - tw) // 2, canvas_h - 60), brand_text, fill=(100, 80, 60), font=brand_font)

        mockup_dir = product_path.parent
        mockup_path = mockup_dir / f"{product_path.stem}_mockup.jpg"
        canvas.save(str(mockup_path), "JPEG", quality=90)
        logger.info("Mockup → %s", mockup_path)
        return mockup_path

    # ------------------------------------------------------------------
    # Dispatch
    # ------------------------------------------------------------------

    def create(self, concept: ProductConcept) -> dict:
        """Create a product, its mockup, and metadata. Return result dict."""
        method_name = self.CATEGORY_MAP.get(concept.category)
        if not method_name:
            raise ValueError(
                f"Unknown category '{concept.category}'. "
                f"Valid: {', '.join(self.CATEGORY_MAP)}"
            )
        creator = getattr(self, method_name)
        product_path = Path(creator(concept))
        mockup_path = self.generate_mockup(product_path)
        out_dir = product_path.parent
        files = [product_path.name, mockup_path.name]
        metadata_path = self._save_metadata(concept, out_dir, files)
        return {
            "product_path": str(product_path),
            "mockup_path": str(mockup_path),
            "metadata_path": str(metadata_path),
            "concept": concept.__dict__,
        }


# ---------------------------------------------------------------------------
# Self-test suite
# ---------------------------------------------------------------------------


def run_tests() -> bool:
    """Run validation tests. Returns True if all pass."""
    import traceback

    test_base = DEFAULT_OUTPUT_DIR / "_test_run"
    if test_base.exists():
        shutil.rmtree(test_base)
    test_base.mkdir(parents=True, exist_ok=True)

    creator = ProductCreator(output_dir=test_base)
    results: list[tuple[str, bool, str]] = []

    def record(name: str, fn):
        try:
            fn()
            results.append((name, True, ""))
        except Exception as exc:
            results.append((name, False, f"{exc}\n{traceback.format_exc()}"))

    # 1. Concept creation
    def test_concept_creation():
        c = ProductConcept(
            name="Test Product", slug="test-product",
            category="checklist", description="A test product",
            features=["Priority Items"], style="modern", pages=2,
        )
        assert c.name == "Test Product"
        assert c.slug == "test-product"
        assert c.pages == 2
        assert c.style == "modern"

    record("test_concept_creation", test_concept_creation)

    # 2. Planner generation
    def test_planner_generation():
        c = ProductConcept(
            name="Test Planner", slug="test-planner",
            category="printable_planner", description="Weekly planner",
            features=["weekly"], pages=2,
        )
        path = creator.create_printable_planner(c)
        assert path.exists(), f"PDF not found: {path}"
        assert path.stat().st_size > 0, "PDF is empty"

    record("test_planner_generation", test_planner_generation)

    # 3. Checklist generation
    def test_checklist_generation():
        c = ProductConcept(
            name="Test Checklist", slug="test-checklist",
            category="checklist", description="A checklist",
            features=["To Do", "Doing", "Done"], pages=1,
        )
        path = creator.create_checklist(c)
        assert path.exists(), f"PDF not found: {path}"
        assert path.stat().st_size > 0, "PDF is empty"

    record("test_checklist_generation", test_checklist_generation)

    # 4. Wall art generation
    def test_wall_art_generation():
        c = ProductConcept(
            name="Stay Focused", slug="test-wall-art",
            category="wall_art", description="Typographic wall art",
            features=["typographic"],
            dimensions={"width": 800, "height": 1000},
        )
        path = creator.create_wall_art(c)
        assert path.exists(), f"PNG not found: {path}"
        assert path.stat().st_size > 0, "PNG is empty"

    record("test_wall_art_generation", test_wall_art_generation)

    # 5. Mockup generation
    def test_mockup_generation():
        test_img = Image.new("RGB", (100, 100), (255, 0, 0))
        test_img_path = test_base / "test_red.png"
        test_img.save(str(test_img_path))
        mockup_path = creator.generate_mockup(test_img_path)
        assert mockup_path.exists(), f"Mockup not found: {mockup_path}"
        assert mockup_path.suffix == ".jpg"
        assert mockup_path.stat().st_size > 0, "Mockup is empty"

    record("test_mockup_generation", test_mockup_generation)

    # 6. Metadata saving
    def test_metadata_saving():
        c = ProductConcept(
            name="Meta Test", slug="test-meta",
            category="checklist", description="Metadata test",
        )
        out = creator._ensure_output_dir(c)
        meta_path = creator._save_metadata(c, out, ["test.pdf"])
        assert meta_path.exists(), "metadata.json not found"
        data = json.loads(meta_path.read_text())
        for key in ("name", "slug", "category", "files", "created_at"):
            assert key in data, f"Missing key: {key}"

    record("test_metadata_saving", test_metadata_saving)

    # Print results
    print("\n" + "=" * 60)
    print("  PRODUCT CREATOR — TEST RESULTS")
    print("=" * 60)
    all_pass = True
    for name, passed, err in results:
        status = "PASS" if passed else "FAIL"
        icon = "✅" if passed else "❌"
        print(f"  {icon}  {status}  {name}")
        if not passed:
            all_pass = False
            for line in err.strip().split("\n"):
                print(f"         {line}")
    print("=" * 60)
    print(f"  {'ALL TESTS PASSED' if all_pass else 'SOME TESTS FAILED'}")
    print("=" * 60 + "\n")

    # Clean up
    if all_pass:
        shutil.rmtree(test_base, ignore_errors=True)
    return all_pass


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Digital Product Creator — generate printable planners, "
                    "SVGs, spreadsheets, social templates, wall art, resumes, "
                    "and checklists with mockups.",
    )
    parser.add_argument(
        "--create",
        metavar="JSON",
        help="Path to a JSON file or inline JSON string describing a ProductConcept.",
    )
    parser.add_argument(
        "--mockup",
        metavar="PRODUCT_PATH",
        help="Generate a mockup for an existing product file.",
    )
    parser.add_argument(
        "--list-types",
        action="store_true",
        help="List supported product types.",
    )
    parser.add_argument("--output-dir", metavar="DIR", help="Override output directory.")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
    parser.add_argument("--test", action="store_true", help="Run self-test suite.")

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    output_dir = Path(args.output_dir) if args.output_dir else None
    creator = ProductCreator(output_dir=output_dir)

    if args.test:
        success = run_tests()
        sys.exit(0 if success else 1)

    if args.list_types:
        print("\nSupported Product Types:")
        print("-" * 50)
        types = creator.capabilities.get("product_types", {})
        for key, desc in types.items():
            print(f"  {key:<28} {desc}")
        print()
        sys.exit(0)

    if args.mockup:
        p = Path(args.mockup)
        if not p.exists():
            logger.error("File not found: %s", p)
            sys.exit(1)
        result = creator.generate_mockup(p)
        print(f"Mockup created: {result}")
        sys.exit(0)

    if args.create:
        raw = args.create
        if os.path.isfile(raw):
            with open(raw) as fh:
                data = json.load(fh)
        else:
            data = json.loads(raw)
        concept = ProductConcept(**data)
        result = creator.create(concept)
        print(json.dumps(result, indent=2))
        sys.exit(0)

    parser.print_help()
    sys.exit(0)


if __name__ == "__main__":
    main()
